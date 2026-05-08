import openpyxl

wb = openpyxl.load_workbook("Plantilla-modelo.xlsx")
print(f"Sheets: {wb.sheetnames}")

for sheet_name in wb.sheetnames:
    sheet = wb[sheet_name]
    print(f"\nSheet: {sheet_name}")
    # Print first row
    first_row = [cell.value for cell in next(sheet.rows)]
    print(f"Headers: {first_row}")
    # Print first few rows
    print("First 3 rows of data:")
    for i, row in enumerate(sheet.iter_rows(min_row=2, max_row=4, values_only=True)):
        print(row)

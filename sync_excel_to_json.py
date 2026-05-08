"""
sync_excel_to_json.py
---------------------
Lee Plantilla-modelo.xlsx y genera/actualiza un JSON por cada pestaña
en la carpeta de destino, preservando campos existentes del JSON.

Uso:
    python sync_excel_to_json.py
    python sync_excel_to_json.py --excel otra_ruta.xlsx --out otra_carpeta/
    python sync_excel_to_json.py --dry-run   # muestra cambios sin escribir
"""

import argparse
import json
import os
import re
import sys
from pathlib import Path

import openpyxl

# ──────────────────────────────────────────────
# Configuración por defecto
# ──────────────────────────────────────────────
DEFAULT_EXCEL = "Plantilla-modelo.xlsx"
DEFAULT_OUT   = r"apiService\Infraestructura\Datos\config-centros"

# Palabras clave → tipo de sección
TIPO_KEYWORDS = {
    "bombeo":    ["BOMBEO", "ARQUETA"],
    "quimicos":  ["QUÍMICOS", "QUIMICOS", "QUÍMICO", "QUIMICO"],
    "soplante":  ["SOPLANTE"],
    "deposito":  ["DEPÓSITO", "DEPOSITO"],
    "cuadro":    ["CUADRO ELÉCTRICO", "CUADRO ELECTRICO"],
    "filtro":    ["FILTRO"],
}

# ──────────────────────────────────────────────
# Helpers
# ──────────────────────────────────────────────

def clean(text: str) -> str:
    if not text:
        return ""
    return " ".join(str(text).split())          # colapsa saltos de línea y espacios


def infer_tipo(title: str) -> str:
    title_up = title.upper()
    for tipo, keywords in TIPO_KEYWORDS.items():
        if any(k in title_up for k in keywords):
            return tipo
    return "simple"


def extract_fields(text: str) -> list[dict]:
    """
    Detecta campos de medición en el texto de la tarea.
    Reconoce patrones como: '14 A', '50HZ', '0,3BAR', '11%', 'AMPERIOS', etc.
    Devuelve lista de dicts {clave, sufijo}.
    """
    fields = []
    seen   = set()
    upper  = text.upper()

    patterns = [
        (r"\d[\d,\.]*\s*A\b",           "amperios", "A"),
        (r"\d[\d,\.]*\s*HZ\b",          "hz",       "HZ"),
        (r"\d[\d,\.]*\s*BAR\b",         "bar",      "BAR"),
        (r"\d[\d,\.]*\s*%",             "porcentaje","%" ),
        (r"\bAMPERIOS\b",               "amperios", "A"),
        (r"\bHERTIOS\b|\bFRECUENCIA\b", "hz",       "HZ"),
        (r"\bPRESIÓN\b|\bPRESION\b",    "bar",      "BAR"),
        (r"\bPORCENTAJE\b",             "porcentaje","%"),
    ]

    for pattern, clave, sufijo in patterns:
        if re.search(pattern, upper) and clave not in seen:
            fields.append({"clave": clave, "sufijo": sufijo})
            seen.add(clave)

    return fields


def strip_numbering(text: str) -> str:
    """Elimina prefijos tipo '1.1 ', '2.3 ' del inicio de la tarea."""
    return re.sub(r"^\d+[\.\-]\d+\s*", "", text).strip()


def parse_sheet(sheet) -> list[dict]:
    """Parsea una hoja y devuelve lista de secciones con sus tareas."""
    secciones      = []
    current        = None

    for row in sheet.iter_rows(values_only=True):
        col_a = str(row[0]).strip().upper() if row[0] else ""
        col_d = clean(row[3])

        # ── Cabecera de sección: columna A = "REVISADO" ──
        if col_a == "REVISADO" and col_d:
            prefijo_match = re.match(r"(\d+)", col_d)
            prefijo       = int(prefijo_match.group(1)) if prefijo_match else len(secciones) + 1

            current = {
                "titulo":  col_d,
                "tipo":    infer_tipo(col_d),
                "prefijo": prefijo,
                "tareas":  [],
            }
            secciones.append(current)
            continue

        # ── Tarea: columna A vacía + columna D con texto ──
        if current and col_d and col_a != "REVISADO":
            desc   = strip_numbering(col_d)
            tarea  = {"descripcion": desc}
            campos = extract_fields(col_d)
            if campos:
                tarea["campos"] = campos
            current["tareas"].append(tarea)

    return secciones


def load_existing(json_path: Path) -> dict:
    if json_path.exists():
        with open(json_path, encoding="utf-8") as f:
            return json.load(f)
    return {}


def save_json(data: dict, json_path: Path):
    json_path.parent.mkdir(parents=True, exist_ok=True)
    with open(json_path, "w", encoding="utf-8") as f:
        json.dump(data, f, indent=4, ensure_ascii=False)


# ──────────────────────────────────────────────
# Main
# ──────────────────────────────────────────────

def sync(excel_path: str, out_folder: str, dry_run: bool = False):
    excel_path = Path(excel_path)
    out_folder = Path(out_folder)

    if not excel_path.exists():
        print(f"❌  No se encuentra: {excel_path}")
        sys.exit(1)

    wb      = openpyxl.load_workbook(excel_path, data_only=True)
    results = {"ok": [], "skipped": [], "errors": []}

    for sheet_name in wb.sheetnames:
        center = sheet_name.strip()
        try:
            sheet     = wb[sheet_name]
            secciones = parse_sheet(sheet)

            if not secciones:
                results["skipped"].append(center)
                continue

            json_path    = out_folder / f"{center}.json"
            existing     = load_existing(json_path)

            # Preservar campos extra del JSON que no vienen del Excel
            new_data = {k: v for k, v in existing.items()
                        if k not in ("nombre", "secciones")}
            new_data["nombre"]    = center
            new_data["secciones"] = secciones

            if dry_run:
                print(f"\n── DRY RUN: {center} ──")
                print(json.dumps(new_data, indent=4, ensure_ascii=False)[:800], "…")
            else:
                save_json(new_data, json_path)

            results["ok"].append(center)

        except Exception as e:
            results["errors"].append((center, str(e)))

    # ── Resumen ──
    print(f"\n{'[DRY RUN] ' if dry_run else ''}Sincronización completada")
    print(f"  ✅  Generados : {len(results['ok'])}  → {', '.join(results['ok'])}")
    if results["skipped"]:
        print(f"  ⏭   Saltados  : {len(results['skipped'])}  → {', '.join(results['skipped'])}")
    if results["errors"]:
        print(f"  ❌  Errores   : {len(results['errors'])}")
        for name, err in results["errors"]:
            print(f"       {name}: {err}")


if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Sincroniza Excel → JSON para GTM")
    parser.add_argument("--excel",   default=DEFAULT_EXCEL,  help="Ruta al .xlsx")
    parser.add_argument("--out",     default=DEFAULT_OUT,    help="Carpeta de destino para los JSON")
    parser.add_argument("--dry-run", action="store_true",    help="Muestra cambios sin escribir ficheros")
    args = parser.parse_args()

    sync(args.excel, args.out, dry_run=args.dry_run)
